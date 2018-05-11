'''
Created on Mar 20, 2018

@author: Mammon
'''

from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from _tracemalloc import start 
from tkinter.constants import CURRENT
import os

from midiutil.MidiFile import MIDIFile



# {  pitch related actions  }
sharpList = ["C", "C#", "D", "D#", "E", "F", "F#", "G", "G#", "A", "A#", "B"]
flagList = ["C", "Db", "D", "Eb", "E", "F", "Gb", "G", "Ab", "A", "Bb", "B"]
pitchClassToPitchClassNum = {}
pitchClassNumToPitchClass = {}
minorScale = [0, 2, 3, 5, 7, 8, 11]
majorScale = [0, 2, 4, 5, 7, 9, 11]
pentatonicScale = [0, 2, 4, 7, 9]
bluesScale = [0, 2, 3, 4, 7, 9]
wholetoneScale = [0, 2, 4, 6, 8, 10]
def getPitchType(pitchClass):
    if pitchClass.endswith("b"):
        return 1
    return 0
def getPitchName(pitchClassNumToPitchClass, pitchNum, noteType):
    pitchName = pitchClassNumToPitchClass[(pitchNum % 12, noteType)]
    pitchName += str(int(pitchNum / 12) - 1)
    return pitchName 
def getPitchClassAndOctave(pitchName):
    # C-1   D#-1
    # G10
    # C..Z c..z  # b
    # 0..9 -
    i = 0
    octaveSet = [str(i) for i in range(10)]
    octaveSet.append('-')
    while not pitchName[i] in octaveSet:
        i += 1
    return pitchName[:i], int(pitchName[i:])
def getPitchNumRange(pitchClassToPitchClassNum, pitchName):
    pitchNumRange = []
    for i in range(2):
        pitchClass, octave = getPitchClassAndOctave(pitchName[i])
        pitchNumRange.append(pitchClassToPitchClassNum[pitchClass] + (octave + 1) * 12);
        # print(pitchName[i] + " : " + str(ans[i]))
    return pitchNumRange 
def get5and14(pitchClassToPitchClassNum, pitchClassNumToPitchClass, sharpList, flagList):
    for i in range(0, len(sharpList)):
        pitchClassToPitchClassNum.update({flagList[i]: i})
        pitchClassToPitchClassNum.update({sharpList[i]: i})
        # the second argument is note name type (sharp or flat)
        pitchClassNumToPitchClass.update({(i, 0): sharpList[i]})
        pitchClassNumToPitchClass.update({(i, 1): flagList[i]})
def getPitchNumSnap(pitchNum, scaleName, pitchClass):
    if scaleName == "":
        # give up things after "."
        return int(pitchNum)
    elif scaleName == "major":
        # try from moving 0 to moving 11 semitones
        for i in range(0, 12):
            # move down first (-1), then move up(1)
            for j in range(-1, 1, 2):
                currentPitch = int(pitchNum) + j * i
                # the current pitch order for certain pitchClass
                # for example, pitch class A in pitchClass A -> 0
                # pitch class B in pitchClass G -> 4
                pitchClassIndex = pitchClassToPitchClassNum[pitchClass]
                currentPitchOrderInKey12 = ((currentPitch - pitchClassIndex) % 12 + 12) % 12
                if currentPitchOrderInKey12 in majorScale:
                    return currentPitch
    elif scaleName == "minor":
        # try from moving 0 to moving 11 semitones
        for i in range(0, 12):
            # move down first (-1), then move up(1)
            for j in range(-1, 1, 2):
                currentPitch = int(pitchNum) + j * i
                # the current pitch order for certain pitchClass
                # for example, pitch class A in pitchClass A -> 0
                # pitch class B in pitchClass G -> 4
                pitchClassIndex = pitchClassToPitchClassNum[pitchClass]
                currentPitchOrderInKey12 = ((currentPitch - pitchClassIndex) % 12 + 12) % 12
                if currentPitchOrderInKey12 in minorScale:
                    return currentPitch
    elif scaleName == "pentatonic":
        # try from moving 0 to moving 11 semitones
        for i in range(0, 12):
            # move down first (-1), then move up(1)
            for j in range(-1, 1, 2):
                currentPitch = int(pitchNum) + j * i
                # the current pitch order for certain pitchClass
                # for example, pitch class A in pitchClass A -> 0
                # pitch class B in pitchClass G -> 4
                pitchClassIndex = pitchClassToPitchClassNum[pitchClass]
                currentPitchOrderInKey12 = ((currentPitch - pitchClassIndex) % 12 + 12) % 12
                if currentPitchOrderInKey12 in pentatonicScale:
                    return currentPitch
    elif scaleName == 'wholetone': 
        # try from moving 0 to moving 11 semitones
        for i in range(0, 12):
            # move down first (-1), then move up(1)
            for j in range(-1, 1, 2):
                currentPitch = int(pitchNum) + j * i
                # the current pitch order for certain pitchClass
                # for example, pitch class A in pitchClass A -> 0
                # pitch class B in pitchClass G -> 4
                pitchClassIndex = pitchClassToPitchClassNum[pitchClass]
                currentPitchOrderInKey12 = ((currentPitch - pitchClassIndex) % 12 + 12) % 12
                if currentPitchOrderInKey12 in wholetoneScale:
                    return currentPitch
    elif scaleName == 'blues':
        # try from moving 0 to moving 11 semitones
        for i in range(0, 12):
            # move down first (-1), then move up(1)
            for j in range(-1, 1, 2):
                currentPitch = int(pitchNum) + j * i
                # the current pitch order for certain pitchClass
                # for example, pitch class A in pitchClass A -> 0
                # pitch class B in pitchClass G -> 4
                pitchClassIndex = pitchClassToPitchClassNum[pitchClass]
                currentPitchOrderInKey12 = ((currentPitch - pitchClassIndex) % 12 + 12) % 12
                if currentPitchOrderInKey12 in bluesScale:
                    return currentPitch
    return int(pitchNum)
def getPitchNumListSnap(pitchNumList, pitchClass):
    getPitchNumListSnapes = []
    pitchType = getPitchType(pitchClass)
    for i in range(0, len(pitchNumList)):
        getPitchNumListSnapes.append(getPitchName(pitchClassNumToPitchClass, pitchNumList[i], pitchType))
    return getPitchNumListSnapes



# { Initialization Actions }
def getHierarchyToHierarchySizeList(hierarchyToHierarchySizeList, hierarchyToHierarchyAppearList):
    hierarchyToHierarchySizeList.append(1)
    for i in range(1, len(hierarchyToHierarchyAppearList)):
        hierarchyToHierarchySizeList.append(hierarchyToHierarchySizeList[i - 1] * hierarchyToHierarchyAppearList[i - 1] + 1)


def adjustRange(value, minValue, maxValue, minRange, maxRange):
    return (value - minValue) * (maxRange - minRange) / (maxValue - minValue) + minRange


#return the pitch of a node in a certain hierarchy which has axisValue within [minValue, maxValue]
def getPitch(hierarchy, minValue, maxValue, axisValue, hierarchyToDurationList, keySystem, key):
    range = getPitchNumRange(pitchClassToPitchClassNum, hierarchyToPichNameRangeList[hierarchy])
    minRange = range[0]
    maxRange = range[1]
    pitchInFloat = adjustRange(axisValue, minValue, maxValue, minRange, maxRange)
    return getPitchNumSnap(pitchInFloat, keySystem, key)

#return the onset of a node in a certain hierarchy based on the onset and duration of the previous node
#if it's the first node of one path, then the onset is based on the previous last onset of this hierarchy + the corresponding duration
def getOnset(previousPartEndTime, hierarchy, earlistOnset, shiftValue, minValue, maxValue, hierarchyToDurationList, elementNameToElementWeight):
    #0..1.5 (every 0.5)
    realShiftValue = int(adjustRange(shiftValue, minValue, maxValue, 0.0, 10.99)) * 0.5 
    print("shift before get:" + str(realShiftValue))
    # first note
    if previousPartEndTime == 0 and hierarchy == 0 and earlistOnset == 0:
        return earlistOnset
    return earlistOnset + realShiftValue

#return duration of this hierarchy
def getDuration(hierarchy):
    return elementNameToDuration[hierarchyToElementNameList[hierarchy]]
    
#return velocity
def getVelocity(hierarchy, minValue, maxValue, value):
    # 0 - 100  1 - 80  2 - 60  3 - 40 5 -20
    return int(adjustRange(value, minValue, maxValue, 20, 127))
    # return 100 - 20 * hierarchy
def getTrack(hierarchy):
    return int(hierarchy)

def updateOnsetForHierarchy(hierarchy, baseOnset, minValue, maxValue, value):
    shift = int(adjustRange(value, minValue, maxValue, 0.0, 10.99)) * 0.5
    print("h: " + str(hierarchy) + " base: " + str(baseOnset) + " shift: " + str(shift) + " final: "  + str(baseOnset + shift))
    return baseOnset + shift 

# [6, 5, 4, 3, 2]
def makeRoute(previousPartEndTime, route, hierarchyAxisMinMatrix, hierarchyAxisMaxMatrix, noteInfoList, scaleType, key, pitchClassNumToPitchClass, esheet):
    print (route[0])
    print (route[1])
    # earlistOnset for each note 
    # for the begining of the route, it will be the earlistOnsetFor this hierarchy
    earlistOnsetMax = 0
    earlistOnset = max(hierarchyToStartingPointList[0], previousPartEndTime)
    route.append([])
    route.append([])
    route.append([])
    route.append([])
    for i in range(len(route[0])):
        hierarchy = route[0][i]
        #print("%s : " % (route[hierarchy]))
        elementName = esheet.cell(row=route[1][i], column=1).value
        pitch = getPitch(hierarchy, hierarchyAxisMinMatrix[hierarchy][0], hierarchyAxisMaxMatrix[hierarchy][0], toFloat(esheet.cell(row=route[1][i],column=3).value), hierarchyToDurationList, scaleType, key)
        velocity = getVelocity(hierarchy, hierarchyAxisMinMatrix[hierarchy][1], hierarchyAxisMaxMatrix[hierarchy][1], toFloat(esheet.cell(row=route[1][i],column=4).value))
        # print("earlistOnset:" + str(earlistOnset))
        onset = getOnset(previousPartEndTime, hierarchy, earlistOnset, toFloat(esheet.cell(row=route[1][i], column=5).value), hierarchyAxisMinMatrix[hierarchy][2], hierarchyAxisMaxMatrix[hierarchy][2], hierarchyToDurationList, elementNameToElementWeight)
        # print("Onset:" + str(onset))
        duration = getDuration(hierarchy)
        track = getTrack(hierarchy)
        pitchName = getPitchName(pitchClassNumToPitchClass, pitch, getPitchType(key))
        noteInfoList.append([pitch, velocity, onset, duration, pitchName, track])
        # update earlist onset for current hierarchy (especially hierarchy 0)
        earlistOnset = onset + duration
        if earlistOnset > earlistOnsetMax:
            earlistOnsetMax = earlistOnset
        hierarchyToStartingPointList[hierarchy] = updateOnsetForHierarchy(hierarchy, earlistOnset, hierarchyAxisMinMatrix[hierarchy][2], hierarchyAxisMaxMatrix[hierarchy][2], toFloat(esheet.cell(row=route[1][i], column=5).value))
        hierarchyToStartingPointList[hierarchy] = earlistOnset
        # print("earlistOnset:" + str(earlistOnset))
        # print("hierarchy:" + str(hierarchy) + " hierarchy starting: " + str(hierarchyToStartingPointList[hierarchy]))
        print("P:" + str(getPitchName(pitchClassNumToPitchClass, pitch, getPitchType(key))) + "  O: " + str(onset) + " D: " + str(duration) + " T: " + str(track+1))
        route[2].append(elementName)
        route[3].append(track)
        route[4].append(pitchName)
        route[5].append(duration)
    return earlistOnsetMax

            

def getNoteEventList(noteInfoList, hierarchyMax):
    totalContent = []
    strContent = []
    for i in range(0, len(noteInfoList)):
        if noteInfoList[i][3] == 0: continue
        # 1: Note On Event
        # 0: Note Off Event
        totalContent.append([noteInfoList[i][2], 1, noteInfoList[i][5], noteInfoList[i][0], noteInfoList[i][1], noteInfoList[i][3]])
        totalContent.append([noteInfoList[i][2] + noteInfoList[i][3], 0, noteInfoList[i][5], noteInfoList[i][0], noteInfoList[i][1], noteInfoList[i][3]])
    totalContent = sorted(totalContent, key=lambda x:(x[0], x[1]))
    # print(totalContent)
    return totalContent



def generateRoute(routeList, searchChoiceList, searchStart, hierarchyMax):
    hierarchyToRowIndexList = [0] * len(searchChoiceList)
    hierarchyList = [0, 1, 2, 3, 4, 5]
    for i in range(0, len(searchChoiceList)):
        previousRowIndex = 0
        if i == 0:
            # imaginary value for the choice after the last choice
            # because the node hierarchyToRowIndexList added by at least 1 comparing to the previous choice
            # so subtract searchStart value by 1, and add this 1 back later (hierarchyToRowIndexList[i] = previousRowIndex + 1 + ...)
            previousRowIndex = int(searchStart) - 1
        else:
            previousRowIndex = hierarchyToRowIndexList[hierarchyMax - i]
        hierarchyToRowIndexList[hierarchyMax-1-i]=previousRowIndex+1+int(hierarchyToHierarchySizeList[hierarchyMax-1-i])*int(searchChoiceList[i])
    #skip fake I
    if searchChoiceList[0] == 2:
        hierarchyToRowIndexList.pop()
        hierarchyToRowIndexList.insert(3, specialSiRowIndexList[searchChoiceList[3]])
        hierarchyList = [0, 1, 2, 3, 3, 4]
    print (searchChoiceList)
    routeList.append([hierarchyList, hierarchyToRowIndexList])
    return routeList[len(routeList)-1]

def generateMinMax(hierarchyAxisMinMatrix, hierarchyAxisMaxMatrix, searchStart, searchChoiceList, hierarchyMax, esheet):
    hierarchyToRowIndexList = [0] * len(searchChoiceList)
    for i in range(0, len(searchChoiceList)):
        # previous value
        previousRowIndex = 0
        if i == 0:
            # imaginary value for the choice after the last choice
            # because the node added by at least 1 comparing to the previous choice
            # so subtract searchStart value by 1, and add this 1 back later (hierarchyToRowIndexList[i] = previousRowIndex + 1 + ...)
            previousRowIndex = int(searchStart) - 1
        else:
            previousRowIndex = hierarchyToRowIndexList[hierarchyMax - i]
        hierarchyToRowIndexList[hierarchyMax-1-i]=previousRowIndex+1+int(hierarchyToHierarchySizeList[hierarchyMax-1-i])*int(searchChoiceList[i])
    #skip fake I
    if searchChoiceList[0] == 2:
        hierarchyToRowIndexList.pop()
    for i in range(len(hierarchyToRowIndexList)):
        for j in range(3,6):
            currentValue = toFloat(esheet.cell(row=hierarchyToRowIndexList[i], column=j).value)
            if currentValue < hierarchyAxisMinMatrix[i][j-3]:
                hierarchyAxisMinMatrix[i][j-3] = currentValue
            if currentValue> hierarchyAxisMaxMatrix[i][j-3]:
                hierarchyAxisMaxMatrix[i][j-3] = currentValue
    return hierarchyAxisMinMatrix, hierarchyAxisMaxMatrix 

# search for each permutation of route choice
# esheet: axis value
# minValue, maxValue: minValue and maxValue from column 3-6 for all axis values in esheet (excel sheet)
# pitchContent/onsetContent/durationContent: stores notes' information for generating txt files for reaper js plugin
# return next starting point 
def search(previousPartEndTime, hierarchy, hierarchyMax, searchStart, searchChoiceList, hierarchyAxisMinMatrix, hierarchyAxisMaxMatrix, routeList, noteInfoList, scaleType, key, pitchClassNumToPitchClass, esheet):
    # stop recursion
    if hierarchy == -1:
        # print(searchChoiceList)
        route = generateRoute(routeList, searchChoiceList, searchStart, hierarchyMax)
        return makeRoute(previousPartEndTime, route, hierarchyAxisMinMatrix, hierarchyAxisMaxMatrix, noteInfoList, scaleType, key, pitchClassNumToPitchClass, esheet)
    startingPointMax = -1
    for i in range(0, hierarchyToHierarchyAppearList[hierarchy]):
        searchChoiceList.append(i)
        now = search(previousPartEndTime, hierarchy - 1, hierarchyMax, searchStart, searchChoiceList, hierarchyAxisMinMatrix, hierarchyAxisMaxMatrix, routeList, noteInfoList, scaleType, key, pitchClassNumToPitchClass, esheet)
        if now > startingPointMax:
            startingPointMax = now
        searchChoiceList.pop()
    return startingPointMax

def searchForMinMax(hierarchyAxisMinMatrix, hierarchyAxisMaxMatrix, hierarchy, hierarchyMax, searchChoiceList, searchStart, esheet):
    # stop recursion
    if hierarchy == -1:
        # print(searchChoiceList)
        generateMinMax(hierarchyAxisMinMatrix, hierarchyAxisMaxMatrix, searchStart, searchChoiceList, hierarchyMax, esheet)
        return
    for i in range(0, hierarchyToHierarchyAppearList[hierarchy]):
        searchChoiceList.append(i)
        searchForMinMax(hierarchyAxisMinMatrix, hierarchyAxisMaxMatrix, hierarchy - 1, hierarchyMax, searchChoiceList, searchStart, esheet)
        searchChoiceList.pop()


#string to float
def toFloat(s):
    ans = ""
    dot = -1
    i = 0
    for i in range(0, len(s)):
        if s[i] == '(': 
            break
        if s[i] == '.':
            dot = i
        ans += s[i]
    result = float(ans)
    if i + 1 < len(s) and int(s[i + 1]) >= 5 and dot > -1:
        if result >= 0:
            result += pow(0.1, i - dot - 1)
        else:
            result -= pow(0.1, i - dot - 1)
    return result


def getNoteInfoTxt(folderName, noteInfoList, noteInfoAttributeList):
    for selectTrack in range(len(noteInfoAttributeList)):
        toTxt(folderName + "/" + noteInfoAttributeList[selectTrack] + "Data.txt", noteInfoList, selectTrack)
    for selectTrack in range(5):
        for selectAttribute in range(len(noteInfoAttributeList)):
            toTrackTxt(folderName + "/" + noteInfoAttributeList[selectAttribute] + "-"  + str(selectTrack+1) + ".txt", noteInfoList, selectAttribute, selectTrack)
        
def getNoteInfoListNonOverlapped(noteInfoList, hierarchyMax):
    lastNotePosition = [[-1 for i in range(128)]] * hierarchyMax 
    # onset, duration
    noteInfoList = sorted(noteInfoList, key=lambda x: (x[2], x[3]))
#     for i in range(len(noteInfoList)):
#         # search the previous index for pitch(noteInfoList[i][0]) in track(noteInfoList[i][5])
#         index = lastNotePosition[noteInfoList[i][5]][noteInfoList[i][0]]
#         # if it's the first note, or end time is greater than the current searchStarting time
#         # cut this note
#         if index != -1 and noteInfoList[index][2] + noteInfoList[index][3] > noteInfoList[i][2]:
#             noteInfoList[index][3] = noteInfoList[i][2] - noteInfoList[index][2]
# #             print("cut-------------------")
# #             print("Current: " + str(i) + " onset: " + str(noteInfoList[i][2]))
# #             print("Previous: " + str(index) + " onset: " + str(noteInfoList[index][2]) + " duration: " + str(noteInfoList[index][3]))
# #             print("end cut-------------------")
#         lastNotePosition[noteInfoList[i][5]][noteInfoList[i][0]] = i
    return noteInfoList
def generateFolder(filePath):
    directory = os.path.dirname(filePath)
    # print(directory)
    if not os.path.exists(directory):
        os.makedirs(directory)
def toTxt(filePath, noteInfoList, index):
    generateFolder(filePath)
    f = open(filePath, 'w')
    strContent = []
    for i in range(0, len(noteInfoList)):
        # print(noteInfoList[i][index])
        strContent.append(str(noteInfoList[i][index]) + "\n")
    f.writelines(strContent)
    f.close()
def toTrackTxt(filePath, noteInfoList, selectAttribute, selectTrack):
    generateFolder(filePath)
    f = open(filePath, 'w')
    strContent = []
    for i in range(0, len(noteInfoList)):
        # print(noteInfoList[i][selectAttribute])
        if noteInfoList[i][5] == selectTrack:
            strContent.append(str(noteInfoList[i][selectAttribute]) + "\n")
    f.writelines(strContent)
    f.close()
def getNoteEventTxt(filePath, totalContent, hierarchyMax):
    error = False
    generateFolder(filePath)
    notesIsOn = []
    for i in range(0, hierarchyMax):
        notesIsOn.append([-1] * 128)
    strContent = []
    for i in range(0, len(totalContent)):
        currentStr = ""
        for j in range(0, len(totalContent[i])):
            if j > 0:
                currentStr = currentStr + " "
            currentStr = currentStr + str(totalContent[i][j])
        currentStr = currentStr + "\n"
        strContent.append(currentStr)
    for i in range(0, len(totalContent)):
        pi = notesIsOn[totalContent[i][2]][totalContent[i][3]]
        if totalContent[i][1] == 1 and pi > -1:
            print("Previous: " + str(pi) + " #: " + str(totalContent[pi][2]) + " pitch: " + str(totalContent[pi][3]) + " onset: " + str(totalContent[pi][0]) + " duration: " + str(totalContent[pi][5]))
            print("Current " + str(i) + " #: " + str(totalContent[i][2]) + " pitch: " + str(totalContent[i][3]) + " onset: " + str(totalContent[i][0]))
            error = True 
        elif totalContent[i][1] == 1:
            notesIsOn[totalContent[i][2]][totalContent[i][3]] = i
        elif totalContent[i][1] == 0:
            notesIsOn[totalContent[i][2]][totalContent[i][3]] = -1
    f = open(filePath, 'w')
    f.writelines(strContent)
    f.close()    
    if error:
        print("Complete with note coverage error!")
    print("complete")
    return error
 

specialCRowIndex = 131
specialHRowIndex = 143
specialHCount = 3
ringStartRowIndex = 132
ringSize = 5
hierarchyRingMax = 2
hierarchyAxisRingMinMatrix = [[float('inf')] * 3] * hierarchyRingMax  
hierarchyAxisRingMaxMatrix = [[float('-inf')] * 3] * hierarchyRingMax 

def generateRingMinMax(hierarchyAxisRingMinMatrix, hierarchyAxisRingMaxMatrix, esheet):
    for i in range(hierarchyRingMax):
        rowIndex = ringStartRowIndex
        for z in range(ringSize + 1):
            for j in range(3,6):
                # print("hierarchy:" + str(i) + "row:" + str(rowIndex+1-i))
                axisValue = toFloat(esheet.cell(row=rowIndex+1-i,column=j).value)
                if axisValue < hierarchyAxisRingMinMatrix[i][j-3]: 
                    hierarchyAxisRingMinMatrix[i][j-3] = axisValue
                if axisValue > hierarchyAxisRingMaxMatrix[i][j-3]: 
                    hierarchyAxisRingMaxMatrix[i][j-3] = axisValue
            rowIndex += 2
    for j in range(3,6):
        # print("hierarchy:" + str(1) + "row:" + str(specialCRowIndex))
        axisValue = toFloat(esheet.cell(row=specialCRowIndex, column=j).value)
        if axisValue < hierarchyAxisRingMinMatrix[1][j-3]: 
            hierarchyAxisRingMinMatrix[1][j-3] = axisValue
        if axisValue > hierarchyAxisRingMaxMatrix[1][j-3]: 
            hierarchyAxisRingMaxMatrix[1][j-3] = axisValue
        for i in range(3): 
            # print("hierarchy:" + str(0) + "row:" + str(specialHRowIndex + i))
            axisValue = toFloat(esheet.cell(row=specialHRowIndex+i, column=j).value)
            if axisValue < hierarchyAxisRingMinMatrix[0][j-3]: 
                hierarchyAxisRingMinMatrix[0][j-3] = axisValue
            if axisValue > hierarchyAxisRingMaxMatrix[0][j-3]: 
                hierarchyAxisRingMaxMatrix[0][j-3] = axisValue

def generateRing(routeList, firstPartEndTime, noteInfoList, hierarchyAxisRingMinMatrix, hierarchyAxisRingMaxMatrix, scaleType, key, pitchClassNumToPitchClass, esheet):
    nextC = ringStartRowIndex
    startC = -1
    endTime = 0
    cNum = 12 
    for i in range(ringSize):
        cList = []
        for j in range(cNum):
            cList.append(ringStartRowIndex + ((i+j) % (ringSize+1)) * 2)
        startH = cList[0] + 1
        print([startH] + cList)
        hierarchyList = [0] + [1] * len(cList)
        routeList.append([hierarchyList, [startH] + cList])
        endTimeTemp = makeRoute(firstPartEndTime, routeList[len(routeList)-1], hierarchyAxisRingMinMatrix, hierarchyAxisRingMaxMatrix, noteInfoList, scaleType, key, pitchClassNumToPitchClass, esheet)
        if endTimeTemp > endTime:
            endTime = endTimeTemp
    for i in range(specialHCount):
        routeList.append([[0,1,1], [specialHRowIndex+i, specialCRowIndex, ringStartRowIndex+ringSize*2]])
        endTimeTemp = makeRoute(firstPartEndTime, routeList[len(routeList)-1], hierarchyAxisRingMinMatrix, hierarchyAxisRingMaxMatrix, noteInfoList, scaleType, key, pitchClassNumToPitchClass, esheet)
        if endTimeTemp > endTime:
            endTime = endTimeTemp
    return endTime

def genRouteFile(folderName, dataFileName, routeList):
    filePath = folderName + "/" + dataFileName + "-route.txt";
    routeListFile = open(filePath, 'w')
    fileContent = []
    for routeId in range(len(routeList)):
        route = routeList[routeId]
        for routeItemId in range(2, len(route)):
            fileContent.append(str(route[routeItemId]) + " ")
        fileContent.append("\n")
    routeListFile.writelines(fileContent)
    routeListFile.close()


def genMidiFile(folderName, dataFileName, noteInfoList, hierarchyToTrackNameList, hierarchyMax):
    programChangeList = [40, 0, 47, 42, 43, 58]
    path = folderName + "/" + dataFileName + ".mid"
    midiFile = MIDIFile(hierarchyMax)
    startTime = 0
    channel = 0
    for trackId in range(hierarchyMax):
        midiFile.addTrackName(trackId, startTime, hierarchyToTrackNameList[trackId])
        midiFile.addProgramChange(trackId, channel, startTime, programChangeList[trackId])
        midiFile.addTimeSignature(trackId, startTime, 4, 2, 8, 8)
    for noteId in range(len(noteInfoList)):
        noteInfo = noteInfoList[noteId]
        pitch = noteInfo[0]
        velocity = noteInfo[1]
        onset = noteInfo[2] / 4.0 # in beat
        duration = noteInfo[3] / 4.0 # in beat
        track = noteInfo[5]
        print("pitch:" + str(pitch) + "velocity:" + str(velocity) + "onset:" + str(onset) + "duration:" + str(duration) + "track:" + str(track))
        if duration != 0:
            midiFile.addNote(track, channel, pitch, onset, duration, velocity)
    with open(path, 'wb') as outFile:
        midiFile.writeFile(outFile)
    with open(dataFileName + ".mid", 'wb') as outFile:
        midiFile.writeFile(outFile)
    
        



# get input


# modify arguments
#######################
hierarchyMax = 6
# previous rowMax = 83;
efile=load_workbook(filename = 'dataPython.xlsx')
esheet=efile['Sonification']
elementNameToElementWeight = {"H": 1, "C": 6, "Si": 14, "In": 49, "I": 53}
elementNameToDuration = {"H": 1, "C": 2, "Si": 4, "In": 8, "I": 16}  #1-16th notes
hierarchyToDurationList = [1, 2, 4, 8, 16, 32]
hierarchyToTrackNameList = ["Violin", "Grand Piano", "Percussion", "Cello", "Bass", "Tuba"]
# hierarchyToPichNameRangeList = [['C4', 'C6'], ['C3', 'G4'], ['C4', 'C5'], ['E3', 'C4'], ['E2', 'E3']]
hierarchyToPichNameRangeList = [['C4', 'C6'], ['C2', 'C6'], ['C2', 'C5'], ['C2', 'C4'], ['E2', 'E3'], ['C1', 'C3']]
hierarchyToElementNameList = ["H", "C", "Si", "C", "In", "I"]
# hierarchyToHierarchyAppearList = [3, 3, 3, 1, 1, 3]
# Special Start
specialSiRowIndexList = [128, 129, 130]
specialInIndex = 2; # 0 and 1, single Si (hierarchy 2), to I (hierarchy 5); 2, double Si and no I
# Special End
hierarchyToHierarchyAppearList = [3, 3, 3, 1, 1, 3]
hierarchyToHierarchySizeList = []
hierarchyToStartingPointList = [0, 0, 0, 0, 0, 0]
hierarchyAxisMinMatrix = [[float('inf')] * 3] * hierarchyMax  
hierarchyAxisMaxMatrix = [[float('-inf')] * 3] * hierarchyMax 
noteInfoList = []   
routeList = [] #route => routeHierarchyList; routeIndexList
noteInfoAttributeList = ["pitch", "velocity", "onset", "duration", "pitchName", "track"]
noteEventList = [] 
noteEventAttributeList = ["onset", "onOff", "track", "pitch", "velocity"]
##########################################

folderName = input('Please input the folder name (default: tempData): ')
print (folderName)
if folderName == '':
    folderName = 'tempData'
dataFileName = input('Please input the data file name (default: same as above): ')
if dataFileName == '':
    dataFileName = folderName 
scaleType = input('Please input the scale type (pentatonic or 12tone or tonal or wholetone or bluesScale) (default: pentatonic): ')
if (scaleType == ''):
    scaleType = 'pentatonic'
key = ''
if scaleType == 'pentatonic' or scaleType == 'tonal' or scaleType == 'wholetone' or scaleType == 'bluesScale':
    key = input('Please input the key pentatonic or 12tone or tonal or wholetone or bluesScale) (default: C: ')
if key == '':
    key = 'C' 

get5and14(pitchClassToPitchClassNum, pitchClassNumToPitchClass, sharpList, flagList) 
getHierarchyToHierarchySizeList(hierarchyToHierarchySizeList, hierarchyToHierarchyAppearList)
searchForMinMax(hierarchyAxisMinMatrix, hierarchyAxisMaxMatrix, hierarchyMax - 1, hierarchyMax, [], 2, esheet)

for i in range(len(specialSiRowIndexList)):
    for j in range(3,6):
        currentValue = toFloat(esheet.cell(row=specialSiRowIndexList[i], column=j).value)
        if currentValue < hierarchyAxisMinMatrix[4][j-3]:
            hierarchyAxisMinMatrix[4][j-3] = currentValue
        if currentValue> hierarchyAxisMaxMatrix[4][j-3]:
            hierarchyAxisMaxMatrix[4][j-3] = currentValue

generateRingMinMax(hierarchyAxisRingMinMatrix, hierarchyAxisRingMaxMatrix, esheet)
previousEndTime = generateRing(routeList, 0, noteInfoList, hierarchyAxisRingMinMatrix, hierarchyAxisRingMaxMatrix, scaleType, key, pitchClassNumToPitchClass, esheet)
endTime = search(previousEndTime + 16, hierarchyMax - 1, hierarchyMax, 2, [], hierarchyAxisMinMatrix, hierarchyAxisMaxMatrix, routeList, noteInfoList, scaleType, key, pitchClassNumToPitchClass, esheet)
noteInfoList = getNoteInfoListNonOverlapped(noteInfoList, hierarchyMax) 
getNoteInfoTxt(folderName, noteInfoList, noteInfoAttributeList)
getNoteEventTxt(folderName + "/" + dataFileName + ".txt", getNoteEventList(noteInfoList, hierarchyMax), hierarchyMax)

# generate MIDI
genMidiFile(folderName, dataFileName, noteInfoList, hierarchyToTrackNameList, hierarchyMax)
genRouteFile(folderName, dataFileName, routeList)


print(endTime)