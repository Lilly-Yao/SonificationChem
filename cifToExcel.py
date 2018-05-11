'''
Created on Mar 17, 2018

@author: Mammon
'''

from CifFile import ReadCif
cf = ReadCif("molecule.cif")
my_data = cf.first_block()




#print(my_data)
print(my_data)

columns = ["_atom_site_label", "_atom_site_type_symbol",
            "_atom_site_fract_x", "_atom_site_fract_y", "_atom_site_fract_z"]


#store the loop inside excel
from openpyxl import Workbook
wb = Workbook()

#get current active sheet
ws = wb.active
#change title of the sheet
ws.title = "Sonification"


#change the background color of the tab hoding this title
ws.sheet_properties.tabColor = "1072BA"
for i in range(1, 6):
    ws.cell(row=1,column=i,value=columns[i-1])

for i in range(2,len(my_data["_atom_site_label"])+2):
    for j in range(1,6):
        ws.cell(row=i,column=j,value=my_data[columns[j-1]][i-2])

#wb.save('dataPython.xlsx')        
