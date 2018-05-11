'''
Created on Mar 17, 2018

@author: Mammon
'''

#save a work book

from openpyxl import Workbook
wb = Workbook()

#get current active sheet
ws = wb.active

#add new sheet
ws2 = wb.create_sheet("Sheet2")

#change title of the sheet
ws.title = "Sonification"

ws2.title = "MusicInfo"

#change the background color of the tab hoding this title
ws.sheet_properties.tabColor = "1072BA"
print(wb.sheetnames)
for sheet in wb:
    print(sheet.title)
    
ws['A4'] = 1
print(ws['A4'])

d = ws.cell(row=4,column=2,value=10)
print(ws.cell(row=4,column=2))

wb.save('dataPython.xlsx')